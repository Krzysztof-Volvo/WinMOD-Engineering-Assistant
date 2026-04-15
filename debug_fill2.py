from openpyxl import load_workbook
import pandas as pd
from pathlib import Path

file_path = 'examples/e_stop_rules.xlsx'
df = pd.read_excel(file_path)
print('before', df.to_dict(orient='records'))
print('rule col', 'rule' in df.columns)

wb = load_workbook(file_path, data_only=False)
sheet = wb.active
row1 = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
print('header row values', [cell.value for cell in row1])
print('header types', [cell.data_type for cell in row1])
header_names = [str(cell.value).strip() if cell.value is not None else '' for cell in row1]
print('header_names', header_names)
header_map = {name.lower().replace(' ', '_'): idx for idx, name in enumerate(header_names)}
print('header_map', header_map)
rule_idx = header_map.get('rule')
print('rule_idx', rule_idx)
cell = sheet.cell(row=2, column=rule_idx + 1)
print('cell value', cell.value, 'data_type', cell.data_type, 'repr', repr(cell.value))

index = df.index[0]
print('df index', index, 'value before', df.at[index, 'rule'], pd.isna(df.at[index, 'rule']))
df.at[index, 'rule'] = str(cell.value)
print('value after', df.at[index, 'rule'], type(df.at[index, 'rule']))
print('after df', df.to_dict(orient='records'))
