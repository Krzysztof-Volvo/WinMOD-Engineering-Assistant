from openpyxl import load_workbook
wb = load_workbook('examples/e_stop_rules.xlsx', data_only=False)
ws = wb.active
print('sheet', ws.title)
headers = [cell.value for cell in ws[1]]
print('headers', headers)
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5), start=2):
    values = [cell.value for cell in row]
    data_types = [cell.data_type for cell in row]
    print('row', i, values, data_types)
    for j, cell in enumerate(row, start=1):
        if cell.value is not None and cell.data_type == 'f':
            print(' formula cell', i, j, cell.value)
