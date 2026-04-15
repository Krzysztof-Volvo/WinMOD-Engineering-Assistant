import re
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
import sys

# Main application for PLC signal connector.
# Loads signal, macro, and rule Excel files, then applies matching rules to
# connect macros to the correct PLC signals.
class PLCApp:
    # Initialize the GUI application and set up widgets and internal state.
    def __init__(self, root):
        self.root = root
        self.root.title("PLC Signal Connector")

        self.signals_df = None
        self.macros_df = None
        self.rules_df = None
        self.rule_files = []

        self.top_control_frame = tk.Frame(root)
        self.top_control_frame.pack(pady=8)

        self.load_signals_row = tk.Frame(self.top_control_frame)
        self.load_signals_row.pack(pady=2, anchor="center")
        self.load_signals_desc = tk.Label(self.load_signals_row, text="1) Load PLC Signals Excel", anchor="w")
        self.load_signals_desc.pack(side=tk.LEFT, padx=4)
        self.load_signals_btn = tk.Button(self.load_signals_row, text="Load PLC Signals Excel", command=self.load_signals)
        self.load_signals_btn.pack(side=tk.LEFT)

        self.load_macros_row = tk.Frame(self.top_control_frame)
        self.load_macros_row.pack(pady=2, anchor="center")
        self.load_macros_desc = tk.Label(self.load_macros_row, text="2) Load Macros Excel", anchor="w")
        self.load_macros_desc.pack(side=tk.LEFT, padx=4)
        self.load_macros_btn = tk.Button(self.load_macros_row, text="Load Macros Excel", command=self.load_macros)
        self.load_macros_btn.pack(side=tk.LEFT)

        self.load_rules_row = tk.Frame(self.top_control_frame)
        self.load_rules_row.pack(pady=2, anchor="center")
        self.load_rules_desc = tk.Label(self.load_rules_row, text="3) Load Rules Excel File", anchor="w")
        self.load_rules_desc.pack(side=tk.LEFT, padx=4)
        self.load_rules_btn = tk.Button(self.load_rules_row, text="Load Rules Excel File", command=self.load_rules)
        self.load_rules_btn.pack(side=tk.LEFT)

        self.process_row = tk.Frame(self.top_control_frame)
        self.process_row.pack(pady=6, anchor="center")
        self.process_desc = tk.Label(self.process_row, text="4) Click Process and Connect", anchor="w")
        self.process_desc.pack(side=tk.LEFT, padx=4)
        self.process_btn = tk.Button(self.process_row, text="Process and Connect", command=self.process, state=tk.DISABLED)
        self.process_btn.pack(side=tk.LEFT)

        self.file_label_frame = tk.Frame(root)
        self.file_label_frame.pack(pady=8, fill=tk.X)
        self.signals_label = tk.Label(self.file_label_frame, text="Signals: not loaded")
        self.signals_label.grid(row=0, column=0, sticky="w", padx=4)

        self.macro_listbox = tk.Listbox(root, width=120, height=8)
        self.macro_listbox.pack(pady=5)

        self.device_robot_frame = tk.Frame(root)
        self.device_robot_frame.pack(pady=5, fill=tk.X)

        self.device_frame = tk.Frame(self.device_robot_frame)
        self.device_frame.pack(side=tk.LEFT, padx=5, fill=tk.BOTH, expand=True)
        self.device_list_label = tk.Label(self.device_frame, text="Detected devices:")
        self.device_list_label.pack()
        self.device_listbox = tk.Listbox(self.device_frame, width=60, height=8)
        self.device_listbox.pack(fill=tk.BOTH, expand=True)

        self.robot_frame = tk.Frame(self.device_robot_frame)
        self.robot_frame.pack(side=tk.LEFT, padx=5, fill=tk.BOTH, expand=True)
        self.robot_list_label = tk.Label(self.robot_frame, text="Detected robots:")
        self.robot_list_label.pack()
        self.robot_listbox = tk.Listbox(self.robot_frame, width=60, height=8)
        self.robot_listbox.pack(fill=tk.BOTH, expand=True)

        self.device_signal_frame = tk.Frame(root)
        self.device_signal_frame.pack(pady=5, fill=tk.BOTH, expand=True)
        self.device_signal_label = tk.Label(self.device_signal_frame, text="Select a device to view signals:")
        self.device_signal_label.pack(anchor="w")
        self.device_signal_listbox = tk.Listbox(self.device_signal_frame, width=120, height=8)
        self.device_signal_listbox.pack(fill=tk.BOTH, expand=True)

        self.rule_list_label = tk.Label(root, text="Loaded rules file(s):")
        self.rule_list_label.pack()
        self.rule_listbox = tk.Listbox(root, width=120, height=5)
        self.rule_listbox.pack(pady=5)
        self.refresh_rule_listbox()

        self.device_listbox.bind('<<ListboxSelect>>', self.on_device_selected)

        self.output_text = tk.Text(root, height=16, width=80)
        self.output_text.pack(pady=10)

    # Load the PLC signals Excel file into a DataFrame and update the UI.
    def load_signals(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            self.signals_df = pd.read_excel(file_path)
            self.signals_label.config(text=f"Signals: {file_path}")
            self.output_text.insert(tk.END, f"Loaded signals: {len(self.signals_df)} rows\n")
            self.output_text.insert(tk.END, f"Signal columns: {list(self.signals_df.columns)}\n\n")
            self.refresh_device_and_robot_listboxes()
            self.check_enable_process()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load signals: {e}")

    # Load the macro definitions Excel file and detect the header row automatically.
    def load_macros(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            self.macros_df = read_excel_best_header(
                file_path,
                expected_columns=['Symbol', 'macro signal', 'macro name', 'macro_type', 'macro type']
            )
            self.output_text.insert(tk.END, f"Loaded macros: {len(self.macros_df)} rows\n")
            self.output_text.insert(tk.END, f"Macro columns: {list(self.macros_df.columns)}\n\n")
            self.refresh_macro_listbox()
            self.check_enable_process()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load macros: {e}")

    # Load one or more rule files and merge them into a single rule DataFrame.
    # Rule files can use legacy filter rules or the new macro rule format.
    def load_rules(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_paths:
            return
        new_rules = []
        for path in file_paths:
            try:
                rules_page = read_excel_best_header(
                    path,
                    expected_columns=['field', 'operator', 'value', 'macro signal', 'rule'],
                    require_non_empty_columns=['rule', 'macro signal']
                )
                new_rules.append(rules_page)
                self.rule_files.append(path)
                self.output_text.insert(tk.END, f"Loaded rule file: {path} ({len(rules_page)} rows)\n")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load rule file {path}: {e}")
                return
        if self.rules_df is None:
            self.rules_df = pd.concat(new_rules, ignore_index=True)
        else:
            self.rules_df = pd.concat([self.rules_df] + new_rules, ignore_index=True)
        self.output_text.insert(tk.END, f"Total rules rows: {len(self.rules_df)}\n")
        self.output_text.insert(tk.END, f"Rule columns: {list(self.rules_df.columns)}\n")
        rule_col = get_column_name(self.rules_df, 'rule')
        if rule_col is not None:
            self.output_text.insert(tk.END, f"Loaded rule values: {self.rules_df[rule_col].head(10).tolist()}\n")
        self.output_text.insert(tk.END, "\n")
        self.refresh_rule_listbox()
        self.check_enable_process()

    # Update the rule file list shown in the UI.
    def refresh_rule_listbox(self):
        self.rule_listbox.delete(0, tk.END)
        if self.rule_files:
            for path in self.rule_files:
                self.rule_listbox.insert(tk.END, path)
        else:
            self.rule_listbox.insert(tk.END, "No rule files loaded")

    # Update the device and robot lists shown in the UI from the loaded signal DataFrame.
    def refresh_device_and_robot_listboxes(self):
        self.device_listbox.delete(0, tk.END)
        self.robot_listbox.delete(0, tk.END)
        self.device_signal_listbox.delete(0, tk.END)
        self.device_signal_label.config(text="Select a device to view signals:")
        if self.signals_df is None:
            self.device_listbox.insert(tk.END, "No devices found")
            self.robot_listbox.insert(tk.END, "No robots found")
            return
        devices, robots = extract_devices_and_robots_from_signals(self.signals_df)
        if not devices:
            self.device_listbox.insert(tk.END, "No devices found")
        else:
            for device in devices:
                self.device_listbox.insert(tk.END, device)
        if not robots:
            self.robot_listbox.insert(tk.END, "No robots found")
        else:
            for robot in robots:
                self.robot_listbox.insert(tk.END, robot)

    def on_device_selected(self, event):
        if self.signals_df is None:
            return
        selection = self.device_listbox.curselection()
        if not selection:
            return
        selected_device = self.device_listbox.get(selection[0])
        if selected_device in ["No devices found"]:
            self.device_signal_listbox.delete(0, tk.END)
            return
        self.device_signal_label.config(text=f"Signals for {selected_device}:")
        signals = find_signals_for_device(self.signals_df, selected_device)
        self.device_signal_listbox.delete(0, tk.END)
        if not signals:
            self.device_signal_listbox.insert(tk.END, "No signals found for this device")
            return
        for sig in signals:
            self.device_signal_listbox.insert(tk.END, sig)

    # Update the macro list shown in the UI from the loaded macro DataFrame.
    def refresh_macro_listbox(self):
        self.macro_listbox.delete(0, tk.END)
        if self.macros_df is None:
            return
        for idx, row in self.macros_df.iterrows():
            self.macro_listbox.insert(tk.END, f"{idx + 1}: {row.to_dict()}")

    # Enable the Process button once all required files have been loaded.
    def check_enable_process(self):
        if self.signals_df is not None and self.macros_df is not None and self.rules_df is not None:
            self.process_btn.config(state=tk.NORMAL)

    # Run the matching workflow: apply rules and export matching macro rows.
    def process(self):
        self.output_text.insert(tk.END, "Processing signals and connecting to macros...\n")
        if self.rules_df is None:
            self.output_text.insert(tk.END, "No rules loaded.\n")
            return
        if self.signals_df is None or self.macros_df is None:
            self.output_text.insert(tk.END, "Load all files first.\n")
            return

        if has_macro_rules(self.rules_df):
            self.output_text.insert(tk.END, "Using macro rule format (macro signal + rule).\n")
            matched_macros = apply_macro_rules(
                self.signals_df,
                self.macros_df,
                self.rules_df,
                output_consumer=lambda msg: self.output_text.insert(tk.END, msg)
            )
            if matched_macros.empty:
                self.output_text.insert(tk.END, "No matching macros found using macro rules.\n")
                return
            self.output_text.insert(tk.END, f"Matched macros after applying macro rules: {len(matched_macros)} rows\n")
        else:
            signal_col = get_column_name(self.signals_df, 'name')
            if signal_col is None:
                signal_col = self.signals_df.columns[0]
                self.output_text.insert(tk.END, f"Using first signal column '{signal_col}' for macro matching.\n")
            filtered_signals = apply_signal_rules(
                self.signals_df,
                self.rules_df,
                output_consumer=lambda msg: self.output_text.insert(tk.END, msg)
            )
            if filtered_signals.empty:
                self.output_text.insert(tk.END, "No signals remain after applying rules.\n")
                return
            self.output_text.insert(tk.END, f"Signals after filtering: {len(filtered_signals)} rows\n")
            macro_key_col = find_macro_key_column(self.macros_df)
            if macro_key_col is None:
                self.output_text.insert(tk.END, "No macro key column found in the loaded macros file.\n")
                return
            matched_macros = self.macros_df[self.macros_df[macro_key_col].isin(filtered_signals[signal_col])]
            if matched_macros.empty:
                self.output_text.insert(tk.END, "No matching macros found after filtering.\n")
                return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Text files", "*.txt")], title="Save result as...")
        if not file_path:
            self.output_text.insert(tk.END, "Export cancelled.\n")
            return
        try:
            if file_path.lower().endswith('.xlsx'):
                matched_macros.to_excel(file_path, index=False)
            else:
                matched_macros.to_csv(file_path, sep=",", index=False)
            self.output_text.insert(tk.END, f"Exported {len(matched_macros)} rows to {file_path}\n")
        except Exception as e:
            self.output_text.insert(tk.END, f"Failed to export: {e}\n")



# Apply legacy signal filter rules to the signal DataFrame.
# Returns the subset of signals that match all active rules.
def apply_signal_rules(signals_df, rules_df, output_consumer=None):
    if rules_df is None or rules_df.empty:
        return signals_df
    if not all(col in rules_df.columns for col in ['field', 'operator', 'value']):
        if output_consumer:
            output_consumer("Rule file must contain 'field', 'operator', and 'value' columns.\n")
        return signals_df
    mask = pd.Series(True, index=signals_df.index)
    for _, rule in rules_df.iterrows():
        active = str(rule.get('active', 'yes')).strip().lower()
        if active in ['no', 'false', '0', 'n']:
            continue
        field = str(rule.get('field', '')).strip()
        operator = str(rule.get('operator', 'equals')).strip().lower()
        value = rule.get('value')
        if not field:
            continue
        if field not in signals_df.columns:
            if output_consumer:
                output_consumer(f"Skipping rule because signal column '{field}' is not present.\n")
            continue
        mask &= evaluate_rule(signals_df[field], operator, value)
    return signals_df[mask]


# Evaluate a single filter operation for a pandas Series.
# Supports equality, text matching, numeric comparisons, and regex.
def evaluate_rule(series, operator, value):
    if operator in ['=', '==', 'equals']:
        return series.astype(str) == str(value)
    if operator in ['!=', 'not equals', 'not equal']:
        return series.astype(str) != str(value)
    if operator == 'contains':
        return series.astype(str).str.contains(str(value), na=False)
    if operator == 'startswith':
        return series.astype(str).str.startswith(str(value), na=False)
    if operator == 'endswith':
        return series.astype(str).str.endswith(str(value), na=False)
    if operator == 'in':
        values = [v.strip() for v in str(value).split(',') if v.strip()]
        return series.astype(str).isin(values)
    if operator == 'not in':
        values = [v.strip() for v in str(value).split(',') if v.strip()]
        return ~series.astype(str).isin(values)
    if operator == 'regex':
        return series.astype(str).str.contains(str(value), na=False, regex=True)
    if operator in ['>', 'gt', 'greater than']:
        return pd.to_numeric(series, errors='coerce') > float(value)
    if operator in ['<', 'lt', 'less than']:
        return pd.to_numeric(series, errors='coerce') < float(value)
    if operator in ['>=', 'ge', 'greater or equal', 'greater than or equal']:
        return pd.to_numeric(series, errors='coerce') >= float(value)
    if operator in ['<=', 'le', 'less or equal', 'less than or equal']:
        return pd.to_numeric(series, errors='coerce') <= float(value)
    return series.astype(str) == str(value)


# Evaluate a rule against either a specific signal field or all columns.
def evaluate_rule_on_signals(signals_df, field_name, operator, value):
    if field_name is not None:
        return evaluate_rule(signals_df[field_name], operator, value)
    mask = pd.Series(False, index=signals_df.index)
    for col in signals_df.columns:
        try:
            mask |= evaluate_rule(signals_df[col], operator, value)
        except Exception:
            pass
    return mask


# Normalize header names for case-insensitive lookup.
def normalize_column_name(name):
    return str(name).strip().lower().replace(' ', '_')


# Normalize values for case-insensitive comparison.
def normalize_value(value):
    return str(value).strip().lower()


# Read an Excel file and attempt to find the correct header row.
def read_excel_best_header(file_path, expected_columns=None, require_non_empty_columns=None):
    # Tries the default header row, then later rows if needed.
    require_non_empty_columns = require_non_empty_columns or []
    best_df = pd.read_excel(file_path)
    if expected_columns is None:
        return best_df

    def qualifies(df):
        normalized = [normalize_column_name(c) for c in df.columns]
        if not any(normalize_column_name(col) in normalized for col in expected_columns):
            return False
        for req in require_non_empty_columns:
            col = get_column_name(df, req)
            if col is None:
                return False
            if df[col].dropna().empty:
                return False
        return True

    best_df = fill_missing_formula_values(best_df, file_path, 0)
    if qualifies(best_df):
        return best_df

    for header_row in range(1, 10):
        try:
            df = pd.read_excel(file_path, header=header_row)
        except Exception:
            continue
        df = fill_missing_formula_values(df, file_path, header_row)
        if qualifies(df):
            return df

    return best_df


def fill_missing_formula_values(df, file_path, header_row=0):
    rule_col = get_column_name(df, 'rule')
    if rule_col is None:
        return df
    empty_mask = df[rule_col].isna() | (df[rule_col].astype(str).str.strip() == '')
    if not empty_mask.any():
        return df
    try:
        workbook = load_workbook(file_path, data_only=False)
        sheet = workbook.active
        header_cells = list(sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=False))[0]
        header_names = [str(cell.value).strip() if cell.value is not None else '' for cell in header_cells]
        header_map = {normalize_column_name(name): idx for idx, name in enumerate(header_names)}
        rule_idx = header_map.get(normalize_column_name('rule'))
        if rule_idx is None:
            return df
        if df[rule_col].dtype != object:
            df[rule_col] = df[rule_col].astype(object)
        for df_i in df.index:
            current_value = df.at[df_i, rule_col]
            if pd.isna(current_value) or str(current_value).strip() == '':
                excel_row = header_row + 2 + (df_i if isinstance(df_i, int) else 0)
                cell = sheet.cell(row=excel_row, column=rule_idx + 1)
                value = None
                if getattr(cell, 'data_type', None) == 'f':
                    value = cell.value
                elif cell.value is not None:
                    value = cell.value
                if value is not None and str(value).strip() != '':
                    df.at[df_i, rule_col] = str(value)
    except Exception:
        pass
    return df


# Choose the best candidate column that likely contains macro symbol values.
def find_column_by_header_candidates(df, candidates):
    normalized = {normalize_column_name(col): col for col in df.columns}
    for target in candidates:
        col = normalized.get(normalize_column_name(target))
        if col is not None:
            return col
    return None


def find_all_column_candidates(df, candidates):
    normalized = {normalize_column_name(col): col for col in df.columns}
    found = []
    for target in candidates:
        col = normalized.get(normalize_column_name(target))
        if col is not None and col not in found:
            found.append(col)
    return found


def find_best_symbol_column(df):
    best_col = find_column_by_header_candidates(df, ['Symbol'])
    if best_col is not None and df[best_col].astype(str).str.strip().replace('nan','').ne('').any():
        return best_col

    candidates = []
    for col in df.columns:
        nonempty = df[col].astype(str).str.strip().replace('nan','').ne('')
        count = int(nonempty.sum())
        if count > 0:
            candidates.append((count, col))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


# Find the column used to identify macro rows for rule matching.
def find_macro_key_column(df):
    for candidates in [['macro signal', 'macro_signal'], ['macro name', 'macro_name'], ['Symbol']]:
        for target in candidates:
            col = get_column_name(df, target)
            if col is not None:
                return col
    return find_best_symbol_column(df)


# Find a column name in the DataFrame by normalized target name.
def get_column_name(df, target):
    target_key = normalize_column_name(target)
    for col in df.columns:
        if normalize_column_name(col) == target_key:
            return col
    return None


def get_column_value(row, target):
    target_key = normalize_column_name(target)
    for col in row.index:
        if normalize_column_name(col) == target_key:
            return row[col]
    return None


def is_robot_name(value):
    if not value:
        return False
    last_segment = value.split('+')[-1].strip()
    return bool(re.match(r'^[Rr]\d+$', last_segment))


def extract_devices_and_robots_from_signals(signals_df):
    devices = set()
    robots = set()
    for col in signals_df.select_dtypes(include=['object', 'string']).columns:
        for raw_value in signals_df[col].dropna().astype(str):
            if '=' in raw_value:
                device_name = raw_value.split('=', 1)[0].strip()
                if not device_name:
                    continue
                if is_robot_name(device_name):
                    robots.add(device_name)
                else:
                    devices.add(device_name)
    return sorted(devices), sorted(robots)


def find_signals_for_device(signals_df, device_name):
    results = set()
    if signals_df is None or device_name is None:
        return []
    search_prefix = device_name + '='
    for col in signals_df.select_dtypes(include=['object', 'string']).columns:
        for raw_value in signals_df[col].dropna().astype(str):
            raw_value = raw_value.strip()
            if raw_value.startswith(search_prefix):
                results.add(raw_value)
    return sorted(results)


def extract_plc_tag_from_text(text):
    if not isinstance(text, str):
        return None
    match = re.search(r"(\+[A-Za-z0-9]+(?:\+[A-Za-z0-9]+)+(?:=[^\s]+)?)", text)
    if match:
        return match.group(1)
    return None


def filter_signals_by_macro_channel(signals_df, macro_signal):
    if macro_signal is None or not isinstance(macro_signal, str):
        return signals_df
    channel_match = re.search(r'(?i)ch(\d+)', macro_signal)
    if not channel_match:
        return signals_df
    channel = channel_match.group(1)
    if channel not in ['1', '2']:
        return signals_df
    suffix = f"_{channel}2"
    filtered = []
    for _, row in signals_df.iterrows():
        for val in row:
            if pd.notna(val) and suffix in str(val):
                filtered.append(row)
                break
    if filtered:
        return pd.DataFrame(filtered)
    return signals_df


def resolve_signal_symbol(signal_row, selected_field, rule_text):
    candidates = []
    if selected_field:
        selected_value = get_column_value(signal_row, selected_field)
        if pd.notna(selected_value) and str(selected_value).strip():
            candidates.append(selected_value)

    for preferred in ['Symbol', 'name', 'tag', 'signal', 'Comment', 'comment', 'Description', 'description']:
        preferred_value = get_column_value(signal_row, preferred)
        if pd.notna(preferred_value) and str(preferred_value).strip():
            candidates.append(preferred_value)

    # Try extracting a PLC tag pattern from any signal field.
    for col in signal_row.index:
        val = signal_row.get(col)
        if pd.notna(val):
            tag = extract_plc_tag_from_text(str(val))
            if tag:
                candidates.append(tag)

    for val in candidates:
        if pd.notna(val) and str(val).strip():
            return str(val).strip()

    if isinstance(rule_text, str) and '=' in rule_text:
        rhs = rule_text.split('=', 1)[1].strip()
        if rhs:
            return rhs

    # More robust rule-text matching.
    search = str(rule_text).strip()
    if '=' in search:
        search = search.split('=', 1)[1].strip()
    if search:
        search_norm = search.lower()
        for col in signal_row.index:
            val = signal_row.get(col)
            if pd.notna(val) and search_norm in str(val).lower():
                return str(val).strip()

    for col in signal_row.index:
        val = signal_row.get(col)
        if pd.notna(val) and str(val).strip():
            return str(val).strip()
    return None


def normalize_value(value):
    return str(value).strip().lower()


# Detect whether the rule sheet contains macro-specific rules.
def has_macro_rules(rules_df):
    return (
        rules_df is not None
        and get_column_name(rules_df, 'rule') is not None
        and (
            get_column_name(rules_df, 'macro signal') is not None
            or get_column_name(rules_df, 'macro name') is not None
        )
    )


# Parse a compact rule expression into field/operator/value components.
def parse_rule_text(rule_text, default_field='name', known_fields=None):
    expression = str(rule_text).strip()
    if not expression:
        return default_field, 'equals', ''
    if known_fields is None:
        known_fields = []
    else:
        known_fields = list(known_fields)
    known_keys = {normalize_column_name(f) for f in known_fields}

    def left_is_field(left):
        return normalize_column_name(left) in known_keys

    if expression.startswith('!='):
        return default_field, 'not equals', expression[2:].strip()
    if expression.startswith('=='):
        return default_field, 'equals', expression[2:].strip()
    if expression.startswith('*='):
        return default_field, 'contains', expression[2:].strip()
    for op in ['!=', '==', '=']:
        if op in expression:
            left, right = expression.split(op, 1)
            left = left.strip()
            right = right.strip()
            if op == '!=':
                operator = 'not equals'
            else:
                operator = 'equals'
            if left in ['', '*']:
                if operator == 'equals':
                    return default_field, 'contains', right
                return default_field, operator, right
            if left.endswith('*'):
                return left[:-1], 'contains', right
            if not left_is_field(left):
                return default_field, 'contains', right or left
            return left, operator, right
    if expression.startswith('*') and expression.endswith('*'):
        return default_field, 'contains', expression.strip('*')
    if expression.startswith('*'):
        return default_field, 'endswith', expression[1:]
    if expression.endswith('*'):
        return default_field, 'startswith', expression[:-1]
    return default_field, 'contains', expression


# Apply macro-specific rules and return matching macro rows.
def apply_macro_rules(signals_df, macros_df, rules_df, output_consumer=None):
    macro_key_col = find_macro_key_column(macros_df)
    if macro_key_col is None:
        if output_consumer:
            output_consumer("Macro file must contain a macro key column like 'macro signal' or 'Symbol'.\n")
        return pd.DataFrame()
    if output_consumer:
        output_consumer(f"Using macro key column for rule matching: '{macro_key_col}'\n")

    macro_signal_col = get_column_name(rules_df, 'macro signal')
    rule_col = get_column_name(rules_df, 'rule')
    field_col = get_column_name(rules_df, 'field')
    macro_type_col = get_column_name(rules_df, 'macro type')
    active_col = get_column_name(rules_df, 'active')

    if output_consumer:
        output_consumer(
            f"Macro key column: '{macro_key_col}'.\n"
            f"Rule columns: {list(rules_df.columns)}.\n"
            f"Signal columns: {list(signals_df.columns)}.\n"
        )

    matched_rows = []
    for _, rule in rules_df.iterrows():
        if active_col is not None:
            active = str(rule.get(active_col, 'yes')).strip().lower()
            if active in ['no', 'false', '0', 'n']:
                continue

        macro_signal = rule.get(macro_signal_col)
        if pd.isna(macro_signal) or not str(macro_signal).strip():
            continue
        macro_signal = str(macro_signal).strip()

        rule_text = str(rule.get(rule_col, '')).strip()
        if not rule_text:
            continue

        explicit_field = False
        selected_field = None
        if field_col is not None:
            field_value = rule.get(field_col)
            if not pd.isna(field_value) and str(field_value).strip():
                selected_field = str(field_value).strip()
                explicit_field = True

        selected_field, operator, value = parse_rule_text(
            rule_text,
            default_field=selected_field or 'name',
            known_fields=signals_df.columns
        )
        matched_field = get_column_name(signals_df, selected_field)
        if matched_field is None and explicit_field:
            if output_consumer:
                output_consumer(
                    f"Signal column '{selected_field}' was not found in the signal file. "
                    f"Falling back to search across all signal columns.\n"
                )
            matched_field = None

        signal_mask = evaluate_rule_on_signals(signals_df, matched_field, operator, value)
        matched_signals = signals_df[signal_mask]
        if not matched_signals.empty:
            filtered_signals = filter_signals_by_macro_channel(matched_signals, macro_signal)
            if not filtered_signals.empty:
                matched_signals = filtered_signals
        if matched_signals.empty:
            if output_consumer:
                output_consumer(f"No signals matched rule '{rule_text}' for macro signal '{macro_signal}'.\n")
            continue

        macro_signal_norm = normalize_value(macro_signal)
        key_columns = [macro_key_col] + [c for c in find_all_column_candidates(macros_df, ['macro signal', 'macro_signal', 'macro name', 'macro_name', 'Symbol']) if c != macro_key_col]
        matching_macros = pd.DataFrame()
        used_key_col = None
        for key_col in key_columns:
            macro_symbols = macros_df[key_col].astype(str).apply(normalize_value)
            candidate_matches = macros_df[macro_symbols == macro_signal_norm]
            if not candidate_matches.empty:
                matching_macros = candidate_matches
                used_key_col = key_col
                break
        if matching_macros.empty:
            if output_consumer:
                sample_values = []
                for key_col in key_columns:
                    values = [v for v in macros_df[key_col].astype(str).apply(normalize_value).tolist() if v and v != 'nan']
                    sample_values.append((key_col, values[:20]))
                output_consumer(
                    f"No macros found for macro signal '{macro_signal}'.\n"
                    f"Tried key columns: {key_columns}.\n"
                    f"Available normalized symbols (sample): {sample_values}\n"
                )
            continue
        elif used_key_col != macro_key_col and output_consumer:
            output_consumer(f"Matched macro signal '{macro_signal}' using fallback macro key column '{used_key_col}'.\n")

        macro_name_col = get_column_name(macros_df, 'macro name')
        symbol_col = get_column_name(macros_df, 'Symbol')
        for _, signal_row in matched_signals.iterrows():
            for _, macro_row in matching_macros.iterrows():
                combined = {}
                for col in macros_df.columns:
                    macro_value = macro_row.get(col)
                    signal_value = signal_row.get(col) if col in signal_row.index else None
                    if pd.notna(macro_value) and str(macro_value).strip() != "":
                        combined[col] = macro_value
                    elif signal_value is not None and str(signal_value).strip() != "":
                        combined[col] = signal_value
                    else:
                        combined[col] = macro_value

                if symbol_col is not None and not str(combined.get(symbol_col, '')).strip():
                    best_signal_symbol = resolve_signal_symbol(signal_row, selected_field, rule_text)
                    if pd.notna(best_signal_symbol) and str(best_signal_symbol).strip():
                        combined[symbol_col] = best_signal_symbol

                if macro_name_col is not None:
                    symbol_value = combined.get(symbol_col) if symbol_col is not None else None
                    if not symbol_value:
                        symbol_value = get_column_value(signal_row, 'Symbol')
                    if symbol_value and '=' in str(symbol_value):
                        prefix = str(symbol_value).rsplit('=', 1)[0]
                        if prefix:
                            combined[macro_name_col] = prefix

                macro_signal_col_in_macros = get_column_name(macros_df, 'macro signal')
                if macro_signal_col_in_macros is not None:
                    combined[macro_signal_col_in_macros] = macro_signal

                matched_rows.append(combined)

    if not matched_rows:
        return pd.DataFrame(columns=list(macros_df.columns))
    return pd.DataFrame(matched_rows, columns=list(macros_df.columns))


# Console entry point for command-line processing.
def console_mode(signals_file, macros_file, output_file, rules_file=None):
    signals_df = pd.read_excel(signals_file)
    macros_df = read_excel_best_header(
        macros_file,
        expected_columns=['Symbol', 'macro signal', 'macro name', 'macro_type', 'macro type']
    )
    print(f"Loaded signals: {len(signals_df)} rows")
    print(f"Loaded macros: {len(macros_df)} rows")
    if rules_file:
        rules_df = read_excel_best_header(
            rules_file,
            expected_columns=['field', 'operator', 'value', 'macro signal', 'rule'],
            require_non_empty_columns=['rule', 'macro signal']
        )
        print(f"Loaded macro rules: {len(rules_df)} rows")
        if has_macro_rules(rules_df):
            matched_macros = apply_macro_rules(
                signals_df, macros_df, rules_df,
                output_consumer=lambda msg: print(msg, end='')
            )
            print(f"Matched macros after applying macro rules: {len(matched_macros)} rows")
            matched_macros.to_csv(output_file, sep=",", index=False)
            print(f"Exported {len(matched_macros)} rows to {output_file}")
            return
        signals_df = apply_signal_rules(signals_df, rules_df, output_consumer=lambda msg: print(msg, end=''))
        print(f"Signals after rule filtering: {len(signals_df)} rows")
    signal_col = 'name'
    if signal_col not in signals_df.columns:
        signal_col = signals_df.columns[0]
    macro_key_col = find_macro_key_column(macros_df)
    if macro_key_col is None:
        print("No macro key column found.")
        return
    matched_macros = macros_df[macros_df[macro_key_col].isin(signals_df[signal_col])]
    matched_macros.to_csv(output_file, sep=",", index=False)
    print(f"Exported {len(matched_macros)} rows to {output_file}")


if __name__ == "__main__":
    if len(sys.argv) == 4:
        console_mode(sys.argv[1], sys.argv[2], sys.argv[3])
    elif len(sys.argv) == 5:
        console_mode(sys.argv[1], sys.argv[2], sys.argv[4], rules_file=sys.argv[3])
    else:
        root = tk.Tk()
        app = PLCApp(root)
        root.mainloop()
